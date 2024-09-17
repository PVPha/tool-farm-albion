import pyautogui
import time
import threading
import openpyxl
import threading
import keyboard
import sys
from datetime import datetime
from pynput import mouse
import win32gui
import win32ui
import win32con
import cv2 as cv
import numpy as np

# global variables
stop_listen_mouse = False
position_clicked = ''
key_event_type = ''
list_action = []
bot_action = []
stop_bot_farmer = False
stop_capture_screen = False
# functional
def capture_albion():
    # https://learncodebygaming.com/blog/fast-window-capture
    print('capture')
    global stop_capture_screen
    # get the window size
    hwnd = win32gui.FindWindow(None, 'flower_template - Paint')
    window_rect = win32gui.GetWindowRect(hwnd)
    width = window_rect[2] - window_rect[0]
    height = window_rect[3] - window_rect[1]
    loop_time = time.time()
    while True:
        if stop_capture_screen:
                        break
        # get the window image data
        wDC = win32gui.GetWindowDC(hwnd)
        dcObj = win32ui.CreateDCFromHandle(wDC)
        cDC = dcObj.CreateCompatibleDC()
        dataBitMap = win32ui.CreateBitmap()
        dataBitMap.CreateCompatibleBitmap(dcObj, width, height)
        cDC.SelectObject(dataBitMap)
        cDC.BitBlt((0, 0), (width, height), dcObj, (0, 0), win32con.SRCCOPY)

        # save the image data
        # dataBitMap.SaveBitmapFile(cDC, 'screenshot.bmp')
        
        # convert the raw data into a format opencv can read
        signedIntsArray = dataBitMap.GetBitmapBits(True)
        img = np.frombuffer(signedIntsArray, dtype='uint8')
        img.shape = (height, width, 4)

        # free resource
        dcObj.DeleteDC()
        cDC.DeleteDC()
        win32gui.ReleaseDC(hwnd, wDC)
        win32gui.DeleteObject(dataBitMap.GetHandle())

        # drop the alpha channel to work with opencv matchTemplate()
        img = img[..., :3]

        # make image C_CONTIGUOUS to avoid errors with cv.Rectangle():
        img = np.ascontiguousarray(img)

        # return img
        #display image as video
        # cv.imshow('screenshot', img)

        # print('FPS {}'.format(1 / (time.time() - loop_time)))
        loop_time = time.time()

        # object detection
        threshold = 0.7
        debug_mode='rectangles'
        needle_img = cv.imread('./assets/flower_template_crop.jpg', cv.IMREAD_UNCHANGED)
        needle_w = needle_img.shape[1]
        needle_h = needle_img.shape[0]

        method = cv.TM_CCOEFF_NORMED
        result = cv.matchTemplate(img, needle_img, method)

        locations = np.where(result >= threshold)
        locations = list(zip(*locations[::-1]))

        rectangles = []
        for loc in locations:
            rect = [int(loc[0]), int(loc[1]), needle_w, needle_h]
            rectangles.append(rect)
            rectangles.append(rect)
        rectangles, weights = cv.groupRectangles(rectangles, groupThreshold=1,eps=0.5)
        print(len(rectangles))

        points = []
        if len(rectangles):
            line_color = (0, 255, 0)
            line_type = cv.LINE_4
            marker_color = (255, 0, 255)
            marker_type = cv.MARKER_CROSS

            for(x,y,w,h) in rectangles:
                center_X = x + int(w/2)
                center_Y = y + int(h/2)
                points.append((center_X, center_Y))

                if debug_mode == 'rectangles':
                    top_left = (x,y)
                    bottom_right = (x + w, y + h)

                    cv.rectangle(img, top_left, bottom_right, color=line_color, 
                                lineType=line_type, thickness=2)
                elif debug_mode == 'point':
                    cv.drawMarker(img, (center_X, center_Y), 
                                color=marker_color, markerType=marker_type, 
                                markerSize=40, thickness=2)
        if debug_mode:
            cv.imshow('Matches', img)
            
        if cv.waitKey(1) == ord('q'):
            cv.destroyAllWindows()
            break


threading_capture_albion = threading.Thread(target=capture_albion)

def read_file():
    path = './assets/farm.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for i in range(1, sheet_obj.max_row + 1):
        for j in range(1, sheet_obj.max_column + 1):
            cell_obj = sheet_obj.cell(row = i, column = j)
            print(cell_obj.value, end = ' ')
        print('\n')

def write_file():
    global list_action
    path = './assets/farm.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for i in range(len(bot_action)):
        print(bot_action[i], i)
        if bot_action[i]['type'] == 'keyboard':
            sheet_obj.cell(row = i + 1, column = 1).value = bot_action[i]['key']
            sheet_obj.cell(row = i + 1, column = 2).value = bot_action[i]['event']
            sheet_obj.cell(row = i + 1, column = 3).value = bot_action[i]['time_sleep']
            sheet_obj.cell(row = i + 1, column = 4).value = bot_action[i]['type']
        if bot_action[i]['type'] == 'mouse':
            sheet_obj.cell(row = i + 1, column = 1).value = bot_action[i]['positionX']
            sheet_obj.cell(row = i + 1, column = 2).value = bot_action[i]['positionY']
            sheet_obj.cell(row = i + 1, column = 3).value = bot_action[i]['time_sleep']
            sheet_obj.cell(row = i + 1, column = 4).value = bot_action[i]['type']
    wb_obj.save(path)

def farmer():
    global stop_bot_farmer
    path = './assets/farm.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for i in range(1, sheet_obj.max_row + 1):
        if stop_bot_farmer:
            break
        type = sheet_obj.cell(row = i, column = 4).value
        time_sleep = sheet_obj.cell(row = i, column = 3).value
        if type == 'keyboard':
            key = sheet_obj.cell(row = i, column = 1).value
            if key == 'shift':
                event = sheet_obj.cell(row = i, column = 2).value
                if event == 'down':
                    pyautogui.keyDown('shift')
                if event == 'up':
                    pyautogui.keyUp('shift')
            if key != 'shift':
                pyautogui.press(key)
            time.sleep(time_sleep)
        if type == 'mouse':
            position_x = sheet_obj.cell(row = i, column = 1).value
            position_y = sheet_obj.cell(row = i, column = 2).value
            pyautogui.moveTo(position_x, position_y)
            pyautogui.click()
            time.sleep(time_sleep)
    else:
        print('oh tired farmer')
threading_bot_farmer = threading.Thread(target=farmer)
def detect_mouse_click():
    global stop_listen_mouse
    def on_click(x, y, button, pressed):
        global position_clicked
        global list_action
        global stop_listen_mouse
        if pressed:
            position = (x, y)
            if position_clicked != position:
                position_clicked = position
                time = datetime.now()
                print(f'position {position_clicked}')
                list_action.append({'positionX': x, 'positionY': y, 'time': time, 'type': 'mouse'})
            if button.name == 'right':
                print('right click')
                stop_listen_mouse = True
    mouse_listener = mouse.Listener(on_click=on_click)
    mouse_listener.start()
    if stop_listen_mouse:
        mouse_listener.stop()
threading_detect_mouse_click = threading.Thread(target=detect_mouse_click)

def detect_keyboard():
    global key_event_type
    global stop_listen_mouse
    global bot_action
    global stop_bot_farmer
    global stop_capture_screen
    print(f'm: mouse click \nw: save file \nesc: quit')
    while True:
        key_event = keyboard.read_event(suppress=True)
        if key_event.event_type == keyboard.KEY_UP and key_event.event_type != key_event_type:
            key_event_type = key_event.event_type
            if key_event.name == 'shift':
                print('shift up')
                # time = datetime.now()
                # list_action.append({'key': 'shift','event':'up', 'time': time, 'type': 'keyboard'})
        if key_event.event_type == keyboard.KEY_DOWN and key_event.event_type != key_event_type:
            key_event_type = key_event.event_type
            if key_event.name == 'm':
                if(threading_detect_mouse_click.is_alive() == False):
                    threading_detect_mouse_click.start()
            if key_event.name == 'w':
                print('save file')
                print(list_action)
                length = len(list_action)
                result = []
                for (i, action) in enumerate(list_action):
                    time_sleep = 0
                    if i + 1 < length:
                        time_sleep = (list_action[i + 1]['time'] - action['time']).total_seconds()
                    result.append({**action, 'time_sleep': time_sleep})
                print(result)
                bot_action = result
                write_file()
            if key_event.name == 'f':
                threading_bot_farmer.start()
            if key_event.name == 'shift':
                print('shift down')
                time = datetime.now()
                list_action.append({'key': 'shift','event':'down', 'time': time, 'type': 'keyboard'})
            if key_event.name == 'c':
                if(threading_capture_albion.is_alive() == False):
                    threading_capture_albion.start()
            if key_event.name == 'esc':
                if threading_detect_mouse_click.is_alive():
                    stop_listen_mouse = True
                if threading_capture_albion.is_alive():
                    stop_capture_screen = True
                if threading_bot_farmer.is_alive():
                    stop_bot_farmer = True
                sys.exit(0)
if __name__ == "__main__":
    detect_keyboard()