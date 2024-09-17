import openpyxl

def read_file():
    path = './assets/farm.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    # read 1 row
    # cell_obj = sheet_obj.cell(row = 1, column = 1)
    # print(cell_obj.value, sheet_obj.max_row)

    # read all sheet
    for i in range(1, sheet_obj.max_row + 1):
        for j in range(1, sheet_obj.max_column + 1):
            cell_obj = sheet_obj.cell(row = i, column = j)
            print(cell_obj.value, end = ' ')
        print('\n')
def write_file():
    path = './assets/islands.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    current_row = sheet_obj.max_row
    sheet_obj.cell(row = current_row + 1, column = 2).value = 'hello'
    wb_obj.save(path)
if __name__ == "__main__":
    # write_file()
    read_file()  