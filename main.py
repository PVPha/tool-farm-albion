import keyboard
import time
import pyautogui
import sys
import threading
import math
import openpyxl
from datetime import datetime
from pynput import mouse
# global variables
stop_get_position = False
stop_listen_mouse = False
current_position = ''
measure_average_time = [{'width': 0, 'time': 0}]
list_action = []

def click_target_position():
    path = './assets/farm.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    for i in range(1, max_row + 1):
        target_x = sheet_obj.cell(row = i, column = 1).value
        target_y = sheet_obj.cell(row = i, column = 2).value
        sleep_time = sheet_obj.cell(row = i, column = 3).value
        if target_x is None or target_y is None or sleep_time is None:
            continue
        print(f'target x {target_x} -- target y {target_y} -- sleep time {sleep_time}')
        time.sleep(sleep_time)
        pyautogui.moveTo(target_x, target_y)
        pyautogui.click()
        
    # target_positon = [{'x': 964, 'y': 480, 'time': 0.343589}, {'x': 1291, 'y': 234, 'time': 1.253179}, {'x': 1410, 'y': 238, 'time': 3.247101}, {'x': 1317, 'y': 185, 'time': 4.440492}, {'x': 883, 'y': 149, 'time': 3.675326}, {'x': 1103, 'y': 183, 'time': 5.249549}]
    # for i, target in enumerate(target_positon):
    #     if i == len(target_positon) - 1:
    #         break
    #     target_x = target['x']
    #     target_y = target['y']
    #     # sleep_time = abs(target_x - target_positon[i + 1]['x']) * 0.015
    #     sleep_time = target['time']
    #     time.sleep(sleep_time)
    #     print(f'sleep time {sleep_time}')
    #     pyautogui.moveTo(target_x, target_y)
    #     pyautogui.click()
        
    
def measure_speed(x, y):
    position = []
    position.append({x, y})
    if len(position) > 1:
        print(position)
        print('measure speed')
        time.sleep(0.1)

def auto_press_keyboard(duration=0.1):
    key_press = ['h', 'e', 'l', 'l', 'o', 'space', 'w', 'o', 'r', 'l', 'd']
    try:
        for key in key_press:
            pyautogui.press(key)
            time.sleep(duration)
        else:
            print('quit')
    except:
        print(f'error auto in')

def get_position_mouse():
    global stop_get_position
    global current_position
    while not stop_get_position:
        position_click = pyautogui.position()
        if current_position != position_click:
            current_position = position_click
            print(f'position {position_click}')
get_position_thread = threading.Thread(target=get_position_mouse)

def on_click(x, y, button, pressed):
    global list_action
    global measure_average_time
    if pressed:
        time = datetime.now()
        list_action.append({'x':x, 'y':y, 'time': time, 'type': 'mouse'})
        print(f'x {x} -- y {y} -- time {time}')
        if len(list_action) > 1:
            last_index = len(list_action) - 1
            if list_action[last_index]['type'] != 'mouse' or list_action[last_index - 1]['type'] != 'mouse':
                return
            previous_time = list_action[last_index - 1]['time']
            current_time = list_action[last_index]['time']
            measure_time = current_time - previous_time
            width = 0
            if list_action[last_index - 1]['x'] or list_action[last_index - 1]['y']:
                width = math.sqrt(abs(abs(list_action[last_index]['x'] - list_action[last_index - 1]['x'])** 2 - abs(list_action[last_index]['y'] - list_action[last_index - 1]['y'])** 2))
            print(f'width {width} -- measure time {measure_time.total_seconds()}')
            measure_average_time.append({'width': width, 'time': measure_time.total_seconds()})
        
def detect_mouse_click():
    global stop_listen_mouse
    mouse_listener = mouse.Listener(on_click=on_click)
    mouse_listener.start()
    if stop_listen_mouse:
        mouse_listener.stop()
thread_detect_mouse = threading.Thread(target=detect_mouse_click)

def detect_keypress():
    global stop_get_position
    global current_position
    global stop_listen_mouse
    global list_action
    global measure_average_time
    print('press esc to quit!!!')
    interval = 0
    event = ''
    while True:
        key_event = keyboard.read_event(suppress=True)
        if key_event.event_type == keyboard.KEY_UP and key_event.event_type != event:
            event = key_event.event_type
            key_name = key_event.name
            if key_name == 'shift':
                print('shift up')
                time = datetime.now()
                list_action.append({'key': 'shift','event':'up', 'time': time, 'type': 'keyboard'})
        if key_event.event_type == keyboard.KEY_DOWN and key_event.event_type != event:
            key_name = key_event.name
            event = key_event.event_type
            if key_name == 'esc':
                print('quit')
                stop_get_position = True
                stop_listen_mouse = True
                sys.exit(0)
            print(f"key pressed: {key_name} -- key event {key_event.event_type}")
            if key_name == 's':
                # auto_press_keyboard()
                if(get_position_thread.is_alive() == False):
                    get_position_thread.start()
            if key_name == 'm':
                list_action = [{'x':0, 'y':0, 'time': datetime.now(), 'type': 'mouse'}]
                if(thread_detect_mouse.is_alive() == False):
                    thread_detect_mouse.start()
                # click_target_position()
                # print('movement'
            if key_name == 'shift':
                print('shift down')
                time = datetime.now()
                list_action.append({'key': 'shift','event':'down', 'time': time, 'type': 'keyboard'})
            if key_name == 'r':
                average_time = 0
                if len(measure_average_time) > 0:
                    total_measure = 0
                    for measure in measure_average_time:
                        if measure['width'] == 0 or measure['time'] == 0:
                            continue
                        total_measure += (measure['time'] / measure['width'])
                    average_time = total_measure / len(measure_average_time)
                    print(f'average time {average_time}')
                # save target position
                path = './assets/farm.xlsx'
                wb_obj = openpyxl.load_workbook(path)
                sheet_obj = wb_obj.active
                length = len(list_action)
                print(f'list position {len(list_action)} , measure time {len(measure_average_time)}')
                for i in range(1, length):
                    if list_action[i]['type'] == 'mouse':
                        # target x
                        sheet_obj.cell(row = i, column =1).value = list_action[i]['x']
                        # target y
                        sheet_obj.cell(row = i, column =2).value = list_action[i]['y']
                        # sleep time
                        # sheet_obj.cell(row = i, column =3).value = measure_average_time[i]['width'] * average_time
                        sheet_obj.cell(row = i, column =3).value = measure_average_time[i]['time']
                        # type
                        sheet_obj.cell(row = i, column =4).value = measure_average_time[i]['type']
                    if list_action[i]['type'] == 'keyboard':
                        # key
                        sheet_obj.cell(row = i, column =1).value = list_action[i]['key']
                        # sleep time
                        sheet_obj.cell(row = i, column =2).value = list_action[i]['time']
                        # type
                        sheet_obj.cell(row = i, column =3).value = list_action[i]['type']
                wb_obj.save(path)
            if key_name == 'h':
                print('harvest')
            if key_name == 'p':
                print(f'{pyautogui.position()}')
            if key_name == 'i':
                print(f'interval:{interval}')
                # print(f"Elapsed time: {interval:.2f} seconds")
                interval = datetime.now().time()
            if key_name == 'b':
                click_target_position()

if __name__ == "__main__":
    detect_keypress()
    # get_position_mouse()
    # click_target_position()